import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Reperibilita"

headers = ["DataInizio", "DataFine", "NomePersona", "NumeroInoltro", "EmailPersona", "Note", "Stato"]
header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

widths = [18, 18, 20, 18, 26, 30, 30]
for col_idx, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

date_format = "dd/mm/yyyy hh:mm"

rows = [
    (datetime(2026, 8, 10, 8, 0), datetime(2026, 8, 17, 8, 0), "Mario Rossi", "+39 333 1234567", "mario.rossi@azienda.it", "", ""),
    (datetime(2026, 8, 17, 8, 0), datetime(2026, 8, 24, 8, 0), "Giulia Bianchi", "+39 333 2345678", "giulia.bianchi@azienda.it", "", ""),
    (datetime(2026, 8, 24, 8, 0), datetime(2026, 8, 31, 8, 0), "Luca Verdi", "+39 333 3456789", "luca.verdi@azienda.it", "", ""),
    (datetime(2026, 8, 31, 8, 0), datetime(2026, 9, 7, 8, 0), "Sara Ferrari", "+39 333 4567890", "sara.ferrari@azienda.it", "", ""),
    (datetime(2026, 9, 7, 8, 0), datetime(2026, 9, 14, 8, 0), "Marco Colombo", "+39 333 5678901", "marco.colombo@azienda.it", "", ""),
]

for row_idx, row_data in enumerate(rows, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Arial")
        if col_idx in (1, 2):
            cell.number_format = date_format

ws.freeze_panes = "A2"

# --- Foglio Istruzioni ---
ws2 = wb.create_sheet("Istruzioni")
ws2.column_dimensions["A"].width = 100

title_font = Font(name="Arial", bold=True, size=13)
normal_font = Font(name="Arial", size=11)
bold_font = Font(name="Arial", bold=True, size=11)

lines = [
    ("Calendario Reperibilita - Istruzioni", title_font),
    ("", normal_font),
    ("Questo file viene letto automaticamente dall'app Android di gestione reperibilita' per", normal_font),
    ("determinare chi e' reperibile in ogni momento e impostare di conseguenza l'inoltro delle", normal_font),
    ("chiamate sulla SIM dedicata.", normal_font),
    ("", normal_font),
    ("Colonne del foglio 'Reperibilita':", bold_font),
    ("  A - DataInizio: data e ora di inizio turno (formato gg/mm/aaaa hh:mm)", normal_font),
    ("  B - DataFine: data e ora di fine turno (formato gg/mm/aaaa hh:mm)", normal_font),
    ("  C - NomePersona: nome e cognome della persona reperibile in quel turno", normal_font),
    ("  D - NumeroInoltro: numero di cellulare della persona, in formato internazionale", normal_font),
    ("      (es. +39 333 1234567). E' il numero su cui le chiamate verranno inoltrate.", normal_font),
    ("  E - EmailPersona: email della persona, usata per eventuali notifiche mirate", normal_font),
    ("  F - Note: campo libero facoltativo", normal_font),
    ("  G - Stato: NON compilare a mano. Viene scritta automaticamente dall'app quando", normal_font),
    ("      attiva un turno (es. 'ATTIVATO 2026-08-10T08:00 (Mario Rossi)'), a scopo di log.", normal_font),
    ("", normal_font),
    ("Regole importanti:", bold_font),
    ("  - Una riga per turno.", normal_font),
    ("  - Le date dei turni non devono sovrapporsi: la DataFine di un turno deve coincidere", normal_font),
    ("    o precedere la DataInizio del turno successivo.", normal_font),
    ("  - Se per un certo orario non esiste alcuna riga che lo copre, l'app segnala un errore", normal_font),
    ("    (SMS + mail al numero/indirizzo di servizio configurato nell'app).", normal_font),
    ("  - Il file puo' risiedere in locale sul device oppure su una libreria documenti", normal_font),
    ("    SharePoint, a seconda di come e' configurata l'app nelle Impostazioni.", normal_font),
]

r = 1
for text, font in lines:
    cell = ws2.cell(row=r, column=1, value=text)
    cell.font = font
    r += 1

wb.save("Calendario_Reperibilita.xlsx")
print("OK")
